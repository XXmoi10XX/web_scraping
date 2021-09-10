# 概要
# seleniumを使ってmedia studioにある動画を自動投稿します。

from time import sleep,time
from selenium import webdriver 
import requests 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import pandas as pd
import ffmpeg
import os
import openpyxl
import pprint
import random 
import datetime
import urllib.request, urllib.error
from bs4 import BeautifulSoup
import socks, socket
import subprocess
import random


options = webdriver.ChromeOptions()


# 2. シークレットモードでの使用
options.add_argument('--incognito')
# options.add_argument('--headless')
options.add_argument(
    '--user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.128 Safari/537.36')
# options.add_argument('--proxy-server=socks5://127.0.01:9050')



#step1 : driverを作成する
driver = webdriver.Chrome(
    executable_path='/Users/yoshizawamasaaki/Desktop/lessonn/scraping/tools/chromedriver',
    options=options
)
driver.implicitly_wait(10)



# ーーーーーーーーーーーアフェリエイトリンク取得ーーーーーーーーーーー


wb = openpyxl.load_workbook('/Users/yoshizawamasaaki/Documents/sample_movie/MgsSampleMp4.xltx')

ws = wb.worksheets[0]



movie_info_list = {}
for k, v in zip(ws['A'],ws['B']):
    movie_info_list[k.value] = v.value


# ーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーーー




# --------------ログイン---------------



# アカウント情報
account_inf = ['sadXJH9M5g@gmail.com','kumonoito10','09015809420']
account = random.choice(account_inf)
password = 'Ek*8.k#4ajNE'


# ツイートしたい文字列


# Twitterログイン実行する処理
def login_twitter():
    # ログインページを開く
    driver.get('https://studio.twitter.com/library')
    sleep(2)  # 待ち
    # account入力
    element_account = driver.find_element_by_name("session[username_or_email]")
    element_account.send_keys(account)
    sleep(2)  # 待ち
    # パスワードを入力する
    element_pass = driver.find_element_by_name("session[password]")
    element_pass.send_keys(password)
    sleep(3)  # 動作止める
    # ログインボタンクリック
    element_login = driver.find_element_by_xpath('//*[@data-testid="LoginForm_Login_Button"]')
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    element_login.click()
    sleep(3) # 待ち

# seleniumを起動

# ログイン処理
login_twitter()
sleep(2)

# -----------------------------------


# --------------下にスクロール---------------------

def scroll_down():
    height = driver.execute_script("return document.body.scrollHeight")
    new_height = 0

    while True:
        print(height)
        driver.execute_script(f'window.scrollTo(0, {height});')
        sleep(2)

        new_height = driver.execute_script("return document.body.scrollHeight")
        if height == new_height:
            print('スクロール成功')
            # 最上部へ移動
            driver.execute_script("window.scrollTo(0, 0)")
            break

        height = new_height


scroll_down()

# -----------------------------------

movies = driver.find_elements_by_css_selector('#react-root > div > div.XN9K57Ki._3r9Pt5qv > main > div > div._2RAesh3g._3r9Pt5qv > div._3fNevGDC._3r9Pt5qv > div > div._1_LUvYWq._2d5e4NEa._3r9Pt5qv > div > div > div > div')
# random.shuffle(movies)
count = 0

for i in range(30):

    for movie in movies:
        count += 1

        

        # 動画タイトル取得
        movie_title = movie.find_element_by_css_selector('div:nth-child(2) > div > div.s4mPS2ld > p').text
        print(movie_title)

        tweet_button = movie.find_element_by_css_selector(' div:nth-child(2) > div > div:nth-child(3) > button')
        tweet_button.click()

        sleep(3)

        # 編集ボタン
        #fullScreenScroll > div > div > div > div._2dRIa5go._3r9Pt5qv > div > a
        edit_button = driver.find_element_by_css_selector('div#fullScreenScroll > div > div > div > div._2dRIa5go._3r9Pt5qv > div > a')
        edit_button.click()
        sleep(2)

        # ------------アフェリエイト リンク挿入--------------

     
        # def afe_attach():
        action_link = driver.find_element_by_css_selector('div#tabSettingsPanel > div > div > label:nth-child(5) > input')
                                                    
        if  ".mp4" in movie_title:
                    movie_title = movie_title.replace('.mp4','')
                    print(movie_title)

        for key, value in movie_info_list.items():
            if key == movie_title:
                print('タイトル合致成功')
                action_link.send_keys(value)

        # afe_attach()
        # --------------------------------------------


        # ------------予約投稿--------------

        # def Reserv_Twitter():
            # 予約設定ボタン
        Reserve_button = driver.find_element_by_css_selector('div#fullScreenScroll > div > div > div > div:nth-child(5) > div > div > div._3J6035hs._3r9Pt5qv > button.Button.Button--link._2g7FYFMy > span.Button-label')
        Reserve_button.click()

        sleep(2)

        # div#tabSettingsPanel > div > div > label:nth-child(5) > input

        # 現在の時間と10分後の時間取得
        dt_now = datetime.datetime.now()
        dt2 = datetime.timedelta(minutes= 10 + (30 * count) )
        dt_10leter = dt_now + dt2 
        this_year = dt_now.year

        # this_month = str(dt_now.month) + "月"
        # 予約設定する

        # 10分ごとに予約ツイートするためにfor文で回す

        # 年
        year_select = driver.find_element_by_css_selector('div#fullScreenScroll > div > div > div > div:nth-child(5) > div > div._1Mz0FYWQ._3r9Pt5qv > div._1UApdmW3._3r9Pt5qv > div._247xjXOZ.DatePicker > div.CalendarNavigation > fieldset.FormSelect.CalendarNavigation-yearFormSelect > select')
        year_select.click()

        # year_dec = driver.find_element_by_xpath('//option[text()="{}"]'.format(int(dt_now.year)))
        year_dec = driver.find_element_by_xpath('//option[text()="{}"]'.format(int(dt_10leter.year)))
        
        year_dec.click()
        print("年指定成功")
    

        # 月

        month_select = driver.find_element_by_css_selector('div#fullScreenScroll > div > div > div > div:nth-child(5) > div > div._1Mz0FYWQ._3r9Pt5qv > div._1UApdmW3._3r9Pt5qv > div._247xjXOZ.DatePicker > div.CalendarNavigation > fieldset.FormSelect.CalendarNavigation-monthFormSelect > select')
        month_select.click()

        this_month = int(dt_10leter.month) -1

        month_dec = driver.find_element_by_css_selector('option[value = "{}"]'.format(int(dt_10leter.month) -1))
        month_dec.click()
        print("月指定成功")




        #時間
        hour_minuts = driver.find_element_by_css_selector('div#fullScreenScroll > div > div > div > div:nth-child(5) > div > div._1Mz0FYWQ._3r9Pt5qv > div._1UApdmW3._3r9Pt5qv > div.F9ukX4Vm._3r9Pt5qv > input')
       
        # print(hour_minuts.text)
        # ボックスの中clearにする
       
        # hour_minuts.send_keys(Keys.COMMAND, "a")

        hour_minuts.send_keys(dt_10leter.strftime('%I:%M %p'))
        # hour_minuts.send_keys(Keys.DELETE)
        for j in range(20):
            hour_minuts.send_keys(Keys.LEFT) 
        
        
        for k in range(8):
            hour_minuts.send_keys(Keys.DELETE)
            sleep(2)

        print(dt_10leter.strftime('%H:%M %p'))
       



        # 日

        # dat_button = driver.find_element_by_css_selector('div[data-day="22"] > span')
        # dat_button.click()

        today = dt_now.day 

        dat_button = driver.find_element_by_css_selector('div[data-day="{}"] > span'.format(today))
        dat_button.click()

        sleep(2)

        
    
        # 保存
        save_button = driver.find_element_by_css_selector('div#fullScreenScroll > div > div > div > div:nth-child(5) > div > div._1Mz0FYWQ._3r9Pt5qv > div._1UApdmW3._3r9Pt5qv > div._1JvEh31A._3r9Pt5qv > button.Button.Button--primary.Button--small._1vpXzG5I._1wDu6w9- > span')
        save_button.click()

        sleep(2)




        # 完了
        comp_button =  driver.find_element_by_css_selector('div#fullScreenScroll > div > div > div > div:nth-child(5) > div > div > div._3J6035hs._3r9Pt5qv > button.Button.Button--primary.tweetbtn > span')
        comp_button.click()
        sleep(2)

        print(dt_10leter.strftime('%Y年%m月%d日 %H:%M:%S'))

        # Reserv1qw_Twitter()
    


    # --------------------------------------------


